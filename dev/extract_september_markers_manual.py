#!/usr/bin/env python3
"""
按照load_data.py的逻辑提取September数据
然后填充到EEG_Segmentation.xlsx
对标Aug_05的格式
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import mne
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from collections import Counter

# ============================================================================
# September文件列表 + 对应的Excel行
# ============================================================================

SEPTEMBER_FILES = {
    "sep_12(2).vhdr": {
        "path": r"C:\Users\whyke\github\six-problem\Data\OneDrive_1_01-04-2026\eeg_sep_12(2)\sep_12(2).vhdr",
        "excel_name": "Sep_12(2).xlsx",
        "excel_row": 21,
    },
    "sep_12.vhdr": {
        "path": r"C:\Users\whyke\github\six-problem\Data\OneDrive_1_01-04-2026\eeg_sep_12\sep_12.vhdr",
        "excel_name": "Sep_12.xlsx",
        "excel_row": 22,
    },
    "sep_13(2).vhdr": {
        "path": r"C:\Users\whyke\github\six-problem\Data\OneDrive_1_01-04-2026\eeg_sep_13(2)\sep_13(2).vhdr",
        "excel_name": "Sep_13(2).xlsx",
        "excel_row": 23,
    },
    "sep_13.vhdr": {
        "path": r"C:\Users\whyke\github\six-problem\Data\OneDrive_1_01-04-2026\eeg_sep_13\sep_13.vhdr",
        "excel_name": "Sep_13.xlsx",
        "excel_row": 24,
    },
    "eeg_sep_18.vhdr": {
        "path": r"C:\Users\whyke\github\six-problem\Data\OneDrive_1_01-04-2026\eeg_sep_18\eeg_sep_18.vhdr",
        "excel_name": "Sep_18.xlsx",
        "excel_row": 25,
    }
}

# ============================================================================
# 主函数：提取数据
# ============================================================================

def extract_september_data():
    """按照load_data逻辑提取September数据"""

    print("=" * 100)
    print("SEPTEMBER 数据提取（使用load_data.py逻辑）")
    print("=" * 100)

    results = {}

    for filename, info in SEPTEMBER_FILES.items():
        print(f"\n📂 处理: {filename}")
        print("-" * 100)

        file_path = info["path"]

        # ===== 加载EEG（同load_data.py第16行）=====
        try:
            raw = mne.io.read_raw_brainvision(file_path, preload=True, verbose='ERROR')
        except Exception as e:
            print(f"❌ 无法加载文件: {e}")
            continue

        # ===== RAW INFO（同load_data.py第18-25行）=====
        print(f"\n【RAW INFO】")
        print(f"  Sampling rate: {raw.info['sfreq']} Hz")
        print(f"  Duration: {raw.times[-1]:.2f} sec")
        print(f"  Channels: {len(raw.ch_names)}")

        # ===== ANNOTATIONS（同load_data.py第45-58行）=====
        annotations = raw.annotations
        print(f"\n【ANNOTATIONS】")
        print(f"  Total annotations: {len(annotations)}")

        # 创建DataFrame（同load_data.py第52-56行）
        df_anno = pd.DataFrame({
            'onset_sec': annotations.onset / raw.info['sfreq'],  # 转换为秒
            'duration': annotations.duration,
            'description': annotations.description
        })

        print(f"\n【前20个标记】")
        print(df_anno[['onset_sec', 'description']].head(20).to_string())

        # ===== ANNOTATION COUNTS（同load_data.py第63-64行）=====
        print(f"\n【ANNOTATION COUNTS】")
        counts = df_anno['description'].value_counts()
        print(counts)

        # ===== EVENTS（同load_data.py第69-74行）=====
        events, event_id = mne.events_from_annotations(raw)

        print(f"\n【EVENT ID】")
        print(event_id)
        print(f"Total events: {len(events)}")

        # ===== EVENTS DATAFRAME（同load_data.py第79-88行）=====
        df_events = pd.DataFrame({
            'sample': events[:, 0],
            'event_code': events[:, 2]
        })
        df_events['time_sec'] = df_events['sample'] / raw.info['sfreq']

        # 反向映射label
        inv_map = {v: k for k, v in event_id.items()}
        df_events['label'] = df_events['event_code'].map(inv_map)

        print(f"\n【前30个EVENT】")
        print(df_events[['time_sec', 'label']].head(30).to_string())

        # 保存结果
        results[filename] = {
            'excel_name': info['excel_name'],
            'excel_row': info['excel_row'],
            'total_events': len(annotations),
            'df_anno': df_anno,
            'df_events': df_events,
            'event_id': event_id,
            'annotation_counts': counts.to_dict()
        }

        print(f"\n✅ 提取完成")

    return results

# ============================================================================
# 填充Excel
# ============================================================================

def fill_excel_with_results(results):
    """将提取的数据填充到Excel"""

    print("\n" + "=" * 100)
    print("填充Excel")
    print("=" * 100)

    excel_path = Path(r"C:\Users\whyke\github\six-problem\Data\EEG_Segmentation.xlsx")

    if not excel_path.exists():
        print(f"❌ Excel文件不存在: {excel_path}")
        return

    wb = load_workbook(excel_path)
    ws = wb.active

    for filename, result in results.items():
        row = result['excel_row']
        excel_name = result['excel_name']
        total_events = result['total_events']

        print(f"\n【行{row}: {excel_name}】")
        print(f"  Total events: {total_events}")

        # 填充Name列（通常是列1）
        ws.cell(row, 1).value = excel_name

        # 填充total_event列（列2）
        ws.cell(row, 2).value = total_events

        # 填充match_or_not列（列3）
        match_status = "Match" if total_events > 0 else "NO"
        ws.cell(row, 3).value = match_status

        print(f"  已填充: Name={excel_name}, total_event={total_events}, match={match_status}")

        # 显示标记计数
        print(f"  标记计数: {result['annotation_counts']}")

    # 保存
    wb.save(excel_path)
    print(f"\n✅ Excel已保存: {excel_path}")

# ============================================================================
# 主程序
# ============================================================================

if __name__ == "__main__":

    # 第一步：提取数据
    print("\n【第一步】用load_data逻辑提取September数据\n")
    results = extract_september_data()

    # 第二步：填充Excel
    print("\n【第二步】填充Excel\n")
    fill_excel_with_results(results)

    print("\n" + "=" * 100)
    print("✅ 完成!")
    print("=" * 100)
    print("\n现在请:")
    print("1. 查看输出,对比每个文件的标记")
    print("2. 参考你的handwritten notebook,验证标记号")
    print("3. 手动填入标记号到Excel的各列")
    print("=" * 100)
