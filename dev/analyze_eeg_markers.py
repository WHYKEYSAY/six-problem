#!/usr/bin/env python3
"""
EEG Marker Analysis Script
Analyzes EEG data and markers for six-problem cognitive study
"""

import os
import sys
import json
from pathlib import Path

# Set encoding
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

print("=" * 80)
print("EEG MARKER ANALYSIS SCRIPT")
print("=" * 80)

# Step 1: Try to extract PDF content
print("\n[STEP 1] Extracting content from PDF notebook...")
print("-" * 80)

pdf_path = "Cognitive studies of design_experiment notebook.pdf"
if os.path.exists(pdf_path):
    print(f"[OK] PDF found: {pdf_path}")
    try:
        import pypdf
        reader = pypdf.PdfReader(pdf_path)
        print(f"  Total pages: {len(reader.pages)}")

        # Extract text from first 5 pages
        print("\n  Extracting text from first 5 pages...")
        pdf_content = []
        for i in range(min(5, len(reader.pages))):
            try:
                text = reader.pages[i].extract_text()
                pdf_content.append(f"PAGE {i+1}:\n{text}\n")
            except:
                pass

        if pdf_content:
            print("  [OK] Successfully extracted PDF content")
            with open("pdf_content_extracted.txt", "w", encoding="utf-8") as f:
                f.writelines(pdf_content)
            print("  [OK] Saved to: pdf_content_extracted.txt")
    except ImportError:
        print("  [WARN] pypdf not installed, trying alternative...")
        try:
            import pdfplumber
            with pdfplumber.open(pdf_path) as pdf:
                print(f"  Total pages: {len(pdf.pages)}")
                pdf_content = []
                for i, page in enumerate(pdf.pages[:5]):
                    try:
                        text = page.extract_text()
                        pdf_content.append(f"PAGE {i+1}:\n{text}\n")
                    except:
                        pass
                if pdf_content:
                    print("  [OK] Successfully extracted PDF content")
                    with open("pdf_content_extracted.txt", "w", encoding="utf-8") as f:
                        f.writelines(pdf_content)
        except ImportError:
            print("  [WARN] pdfplumber not installed")
else:
    print(f"[ERROR] PDF not found: {pdf_path}")

# Step 2: Check for available EEG files
print("\n[STEP 2] Searching for EEG data files...")
print("-" * 80)

eeg_extensions = ['.edf', '.set', '.vhdr', '.vmrk', '.eeg']
eeg_files = []

# Search current directory
for ext in eeg_extensions:
    files = list(Path('.').glob(f'**/*{ext}'))
    eeg_files.extend(files)

# Check if D: drive path exists (from load_data.py)
d_drive_path = Path(r"D:\Six problem\six-problem\six-problem\EEG_six_problem")
if d_drive_path.exists():
    print(f"[OK] Found D: drive EEG path: {d_drive_path}")
    for ext in eeg_extensions:
        files = list(d_drive_path.glob(f'**/*{ext}'))
        eeg_files.extend(files)

if eeg_files:
    print(f"[OK] Found {len(eeg_files)} EEG files:")
    for f in eeg_files[:10]:
        print(f"  - {f}")
else:
    print("[ERROR] No EEG files found in current search paths")
    print("  Note: EEG files may be on external drive or different path")

# Step 3: Expected marker definitions
print("\n[STEP 3] Define expected markers from six-problem study...")
print("-" * 80)

expected_markers = {
    'Eye Closed': {
        'code': 'EC',
        'description': 'Eye closed baseline',
        'expected_duration': '30-60 seconds'
    },
    'Problem': {
        'code': 'PROB',
        'description': 'Problem presented/started',
        'expected_count': '~20-30 per session'
    },
    'Solution': {
        'code': 'SOL',
        'description': 'Solution provided/reached',
        'expected_count': '~20-30 per session'
    },
    'Rate': {
        'code': 'RATE',
        'description': 'Rating scale presented',
        'expected_count': '~20-30 per session'
    },
    'Evaluation': {
        'code': 'EVAL',
        'description': 'Evaluation/feedback shown',
        'expected_count': '~20-30 per session'
    },
    'Typing': {
        'code': 'TYPE',
        'description': 'Typing/keyboard input',
        'expected_count': 'Variable'
    }
}

print("Expected marker definitions:")
for marker, info in expected_markers.items():
    print(f"  {marker} ({info['code']}): {info['description']}")

# Step 4: Load existing Excel file structure
print("\n[STEP 4] Checking existing Excel structure...")
print("-" * 80)

try:
    import openpyxl
    wb = openpyxl.load_workbook('EEG_Segmentation.xlsx')
    ws = wb.active

    print(f"[OK] Loaded workbook: {wb.sheetnames}")
    print(f"  Active sheet: {ws.title}")

    # Get header row
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(cell.value)

    print(f"  Columns: {headers}")
    print(f"  Current data rows: {ws.max_row - 1}")

except ImportError:
    print("[WARN] openpyxl not installed")
except Exception as e:
    print(f"[WARN] Error reading Excel: {e}")

# Step 5: Try with MNE to read EEG if available
print("\n[STEP 5] Testing MNE-Python for EEG reading...")
print("-" * 80)

try:
    import mne
    print(f"[OK] MNE-Python version: {mne.__version__}")

    # Try to find and read a sample file
    if eeg_files:
        sample_file = str(eeg_files[0])
        try:
            if sample_file.endswith('.vhdr'):
                raw = mne.io.read_raw_brainvision(sample_file, preload=False)
            elif sample_file.endswith('.edf'):
                raw = mne.io.read_raw_edf(sample_file, preload=False)
            else:
                raw = mne.io.read_raw(sample_file, preload=False)

            print(f"[OK] Successfully loaded: {Path(sample_file).name}")
            print(f"  Channels: {len(raw.ch_names)}")
            print(f"  Sampling rate: {raw.info['sfreq']} Hz")
            print(f"  Duration: {raw.times[-1]:.2f} seconds")
            print(f"  Annotations: {len(raw.annotations)}")

            if len(raw.annotations) > 0:
                print(f"\n  Annotation types:")
                annotations_df = {
                    'onset': raw.annotations.onset,
                    'duration': raw.annotations.duration,
                    'description': raw.annotations.description
                }
                unique_desc = set(annotations_df['description'])
                for desc in sorted(unique_desc):
                    count = sum(1 for d in annotations_df['description'] if d == desc)
                    print(f"    - {desc}: {count}")

        except Exception as e:
            print(f"[WARN] Could not read sample file: {e}")

except ImportError:
    print("[WARN] MNE-Python not installed")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
print("\nNext steps:")
print("1. Review pdf_content_extracted.txt for marker definitions")
print("2. Verify actual markers against expected definitions")
print("3. Process all EEG files to extract event information")
print("4. Fill in EEG_Segmentation.xlsx with findings")
print("=" * 80)
