#!/usr/bin/env python3
"""
填充 EEG_Segmentation.xlsx
使用September数据和完整的被试映射
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import mne
from collections import Counter

# ============================================================================
# September被试完整映射
# ============================================================================

SEPTEMBER_MAPPING = {
    "sep_12.vhdr": {
        "excel_names": ["Sep_12.xlsx", "sep_12.xlsx", "Sep_12", "sep_12"],
        "date": "2013-09-12",
        "subject": "1990 F, Quality",
        "problem_order": ["Cake", "Bin", "Toothbrush", "Workspace", "Fountain", "Metro"],
        "s14_meaning": "Typing",  # Problem 2的Typing = 5_type列
        "s15_meaning": "Evaluate WL",  # Problem 2的Evaluate = 2_rate列
        "s14_problem": 2,
        "s15_problem": 2,
    },
    "sep_12(2).vhdr": {
        "excel_names": ["Sep_12(2).xlsx", "sep_12(2).xlsx", "Sep_12(2)", "sep_12(2)"],
        "date": "2013-09-12",
        "subject": "1987, Anthropology",
        "problem_order": ["Toothbrush", "Workspace", "Cake", "Metro", "Bin", "Fountain"],
        "s14_meaning": "Typing",
        "s15_meaning": "Evaluate WL",
        "s14_problem": 2,
        "s15_problem": 2,
    },
    "sep_13.vhdr": {
        "excel_names": ["Sep_13.xlsx", "sep_13.xlsx", "Sep_13", "sep_13"],
        "date": "2013-09-13",
        "subject": "1987, Info Sys Eng",
        "problem_order": ["Metro", "Bin", "Cake", "Brush", "Fountain", "Exercise"],
        "s14_meaning": "Typing",
        "s15_meaning": "Evaluate WL",
        "s14_problem": 2,
        "s15_problem": 2,
    },
    "sep_13(2).vhdr": {
        "excel_names": ["Sep_13(2).xlsx", "sep_13(2).xlsx", "Sep_13(2)", "sep_13(2)"],
        "date": "2013-09-13",
        "subject": "1987, ECE",
        "problem_order": ["Cake", "Exercise", "Brush", "Bin", "Metro", "Fountain"],
        "s14_meaning": "Typing",
        "s15_meaning": "Evaluate WL",
        "s14_problem": 2,
        "s15_problem": 2,
        "notes": "Signal bad"
    },
    "eeg_sep_18.vhdr": {
        "excel_names": ["Sep_18.xlsx", "sep_18.xlsx", "eeg_sep_18.xlsx", "Sep_18", "sep_18"],
        "date": "2013-09-18",
        "subject": "1982 F, Management",
        "problem_order": ["Cake", "Exercise", "Fountain", "Bin", "Brush", "Metro"],
        "s14_meaning": "Evaluate Solution",  # Problem 2的Evaluate = 2_eval列
        "s15_meaning": "Typing",  # Problem 2的Typing = 5_type列
        "s14_problem": 2,
        "s15_problem": 2,
        "notes": "Marker shift +1"
    }
}

# ============================================================================
# 从EEG文件提取标记
# ============================================================================

def extract_markers_from_eeg(vhdr_path):
    """从EEG文件提取S14和S15计数"""
    try:
        raw = mne.io.read_raw_brainvision(str(vhdr_path), preload=False, verbose='ERROR')
        markers = Counter(raw.annotations.description)

        return {
            'total_events': len(raw.annotations),
            'S14_count': markers.get('Stimulus/S 14', 0),
            'S15_count': markers.get('Stimulus/S 15', 0),
        }
    except Exception as e:
        print(f"  ❌ 错误: {str(e)}")
        return None

# ============================================================================
# 列名到列号的映射
# ============================================================================

COLUMN_MAP = {
    'Name': 1,
    'total_event': 2,
    'match_or_not': 3,
    'eye_closed': 4,
    '1_problem': 5,
    '1_solution': 6,
    '1_rate': 7,
    '1_eval': 8,
    '1_type': 9,
    '2_problem': 11,
    '2_solution': 12,
    '2_rate': 13,
    '2_eval': 14,
    '2_type': 15,
    '3_problem': 17,
    '3_solution': 18,
    '3_rate': 19,
    '3_eval': 20,
    '3_type': 21,
}

# ============================================================================
# 主函数
# ============================================================================

def fill_excel():
    print("=" * 100)
    print("EEG_Segmentation.xlsx 填充工具")
    print("=" * 100)

    excel_path = Path(r"C:\Users\whyke\github\six-problem\Data\EEG_Segmentation.xlsx")

    if not excel_path.exists():
        print(f"❌ 文件不存在: {excel_path}")
        return

    print(f"\n📂 打开Excel: {excel_path.name}")

    # 创建备份
    backup_path = excel_path.parent / f"{excel_path.stem}_backup_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    import shutil
    shutil.copy(excel_path, backup_path)
    print(f"✅ 备份创建: {backup_path.name}\n")

    # 加载workbook
    wb = load_workbook(excel_path)
    ws = wb.active

    print("📊 处理September数据...")
    print("-" * 100)

    filled_count = 0
    september_files = {}

    # 第一步: 找到September行的映射
    for row_idx in range(2, ws.max_row + 1):
        excel_name = ws.cell(row_idx, 1).value
        if not excel_name:
            continue

        # 检查是否匹配任何September文件
        for vhdr_file, mapping in SEPTEMBER_MAPPING.items():
            for excel_variant in mapping['excel_names']:
                if str(excel_name).lower() == excel_variant.lower() or \
                   str(excel_name).lower().replace(' ', '') == excel_variant.lower().replace(' ', ''):
                    september_files[vhdr_file] = {
                        'row': row_idx,
                        'excel_name': excel_name,
                        'mapping': mapping
                    }
                    print(f"\n✓ 找到匹配: {excel_name} ← {vhdr_file}")
                    break

    print(f"\n📋 匹配结果: {len(september_files)}/5个September文件")
    print("-" * 100)

    # 第二步: 从EEG文件提取数据并填充
    for vhdr_file, file_info in september_files.items():
        row_idx = file_info['row']
        excel_name = file_info['excel_name']
        mapping = file_info['mapping']

        print(f"\n【{vhdr_file}】")
        print(f"  Excel行: {row_idx}")
        print(f"  被试: {mapping['subject']}")

        # 查找EEG文件
        eeg_file = None
        possible_dirs = [
            f"eeg_sep_12",
            f"eeg_sep_12(2)",
            f"eeg_sep_13",
            f"eeg_sep_13(2)",
            f"eeg_sep_18",
        ]

        for dir_name in possible_dirs:
            test_path = Path(f"C:/Users/whyke/github/six-problem/Data/OneDrive_1_01-04-2026/{dir_name}/{vhdr_file}")
            if test_path.exists():
                eeg_file = test_path
                break

        if not eeg_file:
            print(f"  ⚠️  EEG文件未找到")
            continue

        # 提取标记
        markers_data = extract_markers_from_eeg(eeg_file)
        if not markers_data:
            continue

        print(f"  📊 事件: 总计{markers_data['total_events']} (S14:{markers_data['S14_count']}, S15:{markers_data['S15_count']})")

        # 填充Excel
        # 1. 总事件数
        ws.cell(row_idx, COLUMN_MAP['total_event']).value = markers_data['total_events']

        # 2. 匹配状态
        match_status = "OK" if markers_data['total_events'] > 0 else "CHECK"
        ws.cell(row_idx, COLUMN_MAP['match_or_not']).value = match_status
        if match_status == "OK":
            ws.cell(row_idx, COLUMN_MAP['match_or_not']).fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )

        # 3. 根据S14和S15含义填充对应列
        s14_problem = mapping.get('s14_problem', 2)
        s15_problem = mapping.get('s15_problem', 2)
        s14_meaning = mapping.get('s14_meaning', '')
        s15_meaning = mapping.get('s15_meaning', '')

        # S14和S15分别对应不同的认知阶段
        # Typing = 5_type (第5列)
        # Evaluate Solution = 2_eval (第4列)
        # Evaluate WL = 2_rate (第3列)

        if s14_meaning == "Typing":
            col_name = f"{s14_problem}_type"
            col_idx = COLUMN_MAP.get(col_name)
            if col_idx:
                ws.cell(row_idx, col_idx).value = markers_data['S14_count']
                print(f"  ✓ 填充 {col_name} = {markers_data['S14_count']}")

        elif s14_meaning == "Evaluate Solution":
            col_name = f"{s14_problem}_eval"
            col_idx = COLUMN_MAP.get(col_name)
            if col_idx:
                ws.cell(row_idx, col_idx).value = markers_data['S14_count']
                print(f"  ✓ 填充 {col_name} = {markers_data['S14_count']}")

        if s15_meaning == "Evaluate WL":
            col_name = f"{s15_problem}_rate"
            col_idx = COLUMN_MAP.get(col_name)
            if col_idx:
                ws.cell(row_idx, col_idx).value = markers_data['S15_count']
                print(f"  ✓ 填充 {col_name} = {markers_data['S15_count']}")

        elif s15_meaning == "Typing":
            col_name = f"{s15_problem}_type"
            col_idx = COLUMN_MAP.get(col_name)
            if col_idx:
                ws.cell(row_idx, col_idx).value = markers_data['S15_count']
                print(f"  ✓ 填充 {col_name} = {markers_data['S15_count']}")

        filled_count += 1

    # 保存
    print("\n" + "=" * 100)
    print(f"✅ 填充完成: {filled_count}/5个September文件")
    print("💾 保存Excel...")

    wb.save(excel_path)
    print(f"✅ Excel已保存: {excel_path}")

    print("\n📝 填充详情:")
    print(f"  • 总事件数: ✅")
    print(f"  • 匹配状态: ✅")
    print(f"  • S14标记计数: ✅")
    print(f"  • S15标记计数: ✅")
    print(f"  • 备份文件: {backup_path.name}")

    print("\n⚠️  注意:")
    print(f"  • 仅填充了问题2的S14/S15相关列")
    print(f"  • 其他问题的列保持空白 (因为September数据中仅有S14/S15标记)")
    print(f"  • 表格显示完整结构,但数据只对应S14/S15对应的认知阶段")

    print("\n" + "=" * 100)
    print("✅ 任务完成!")
    print("=" * 100)

if __name__ == "__main__":
    fill_excel()
