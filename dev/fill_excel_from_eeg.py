#!/usr/bin/env python3
"""
Fill EEG_Segmentation.xlsx from extracted EEG data
Maps extracted events to Excel columns and updates summary table
"""

import os
import sys
import pandas as pd
from pathlib import Path
from datetime import datetime

# Handle encoding
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("[ERROR] openpyxl not installed")
    sys.exit(1)

print("=" * 80)
print("FILL EEG_Segmentation.xlsx FROM EXTRACTED DATA")
print("=" * 80)
print(f"Timestamp: {datetime.now()}\n")

# Step 1: Load extracted events CSV
print("[STEP 1] Loading extracted events...")
print("-" * 80)

csv_file = 'eeg_extracted_events.csv'
if not os.path.exists(csv_file):
    print(f"[ERROR] File not found: {csv_file}")
    print("Please run analyze_eeg_batch.py first to extract EEG markers.")
    sys.exit(1)

try:
    df_events = pd.read_csv(csv_file)
    print(f"[OK] Loaded {len(df_events)} files from {csv_file}")
except Exception as e:
    print(f"[ERROR] Could not read CSV: {e}")
    sys.exit(1)

# Step 2: Load existing Excel file
print("\n[STEP 2] Loading Excel file...")
print("-" * 80)

excel_file = 'EEG_Segmentation.xlsx'
if not os.path.exists(excel_file):
    print(f"[ERROR] File not found: {excel_file}")
    sys.exit(1)

try:
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    print(f"[OK] Loaded workbook: {excel_file}")
    print(f"  Sheet: {ws.title}")
    print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")
except Exception as e:
    print(f"[ERROR] Could not read Excel: {e}")
    sys.exit(1)

# Step 3: Get existing header and structure
print("\n[STEP 3] Analyzing Excel structure...")
print("-" * 80)

# Read header row
headers = []
header_row = 1
for cell in ws[header_row]:
    if cell.value:
        headers.append(cell.value)

print(f"Found {len(headers)} columns:")
print(f"  {headers}")

# Map column indices
header_map = {h: i + 1 for i, h in enumerate(headers)}

# Step 4: Match extracted files to Excel rows
print("\n[STEP 4] Matching extracted files to Excel rows...")
print("-" * 80)

matches = 0
partial_matches = 0
no_matches = 0

# Get existing names from Excel column A (Name column)
excel_names = set()
for row_idx in range(2, ws.max_row + 1):
    cell_value = ws[f'A{row_idx}'].value
    if cell_value:
        excel_names.add(str(cell_value).strip())

print(f"Excel file has {len(excel_names)} file entries")

# Find which extracted files match Excel entries
matched_files = []
for _, row in df_events.iterrows():
    filename = row['File'].strip()
    # Try exact match
    if filename in excel_names:
        matched_files.append((filename, filename, True))
        matches += 1
    else:
        # Try to find partial match (without extension, etc)
        name_base = Path(filename).stem
        partial = None
        for excel_name in excel_names:
            if name_base.lower() in str(excel_name).lower() or str(excel_name).lower() in name_base.lower():
                partial = excel_name
                break

        if partial:
            matched_files.append((filename, partial, False))
            partial_matches += 1
        else:
            no_matches += 1
            matched_files.append((filename, None, False))

print(f"Matched: {matches} exact, {partial_matches} partial, {no_matches} no match")

# Step 5: Update Excel with extracted data
print("\n[STEP 5] Updating Excel with extracted data...")
print("-" * 80)

updated_rows = 0
skipped_rows = 0

for extracted_name, excel_name, is_exact in matched_files:
    if excel_name is None:
        print(f"[WARN] No Excel entry for: {extracted_name}")
        skipped_rows += 1
        continue

    # Find the row in Excel
    row_idx = None
    for r in range(2, ws.max_row + 1):
        if str(ws[f'A{r}'].value).strip() == excel_name:
            row_idx = r
            break

    if row_idx is None:
        print(f"[WARN] Could not find row for: {excel_name}")
        skipped_rows += 1
        continue

    # Get the corresponding extracted row
    extracted_row = df_events[df_events['File'] == extracted_name].iloc[0]

    # Update columns
    try:
        # Total events
        if 'total event' in header_map:
            ws.cell(row_idx, header_map['total event']).value = int(extracted_row.get('Total_Events', 0))

        # Match status
        if 'match or not' in header_map:
            ws.cell(row_idx, header_map['match or not']).value = extracted_row.get('Match_Status', 'UNKNOWN')

        # Event counts for each marker
        marker_columns = [
            'eye_closed',
            '1_problem', '1_solution', '1_rate', '1_eval', '1_type',
            '2_problem', '2_solution', '2_rate', '2_eval', '2_type',
            '3_problem', '3_solution', '3_rate', '3_eval', '3_type',
            '4_problem', '4_solution', '4_rate', '4_eval', '4_type',
            '5_problem', '5_solution', '5_rate', '5_eval', '5_type',
            '6_problem', '6_solution', '6_rate', '6_eval', '6_type',
            'eye_closed_2'
        ]

        for marker in marker_columns:
            if marker in header_map and marker in extracted_row:
                count = extracted_row[marker]
                if pd.notna(count) and count != 'None':
                    ws.cell(row_idx, header_map[marker]).value = int(count)

        updated_rows += 1
        status = "[exact]" if is_exact else "[partial]"
        print(f"[OK] Updated row {row_idx}: {excel_name} {status}")

    except Exception as e:
        print(f"[ERROR] Failed to update row {row_idx}: {e}")
        skipped_rows += 1

# Step 6: Save updated Excel file
print("\n[STEP 6] Saving updated Excel file...")
print("-" * 80)

try:
    # Create backup
    backup_file = f"{excel_file}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    import shutil
    shutil.copy2(excel_file, backup_file)
    print(f"[OK] Created backup: {backup_file}")

    # Save updated file
    wb.save(excel_file)
    print(f"[OK] Saved updated file: {excel_file}")

except Exception as e:
    print(f"[ERROR] Failed to save Excel: {e}")
    sys.exit(1)

# Step 7: Summary
print("\n[STEP 7] Summary")
print("-" * 80)

print(f"Updated rows: {updated_rows}")
print(f"Skipped rows: {skipped_rows}")
print(f"Total processed: {updated_rows + skipped_rows}")

if skipped_rows > 0:
    print(f"\n[WARN] {skipped_rows} rows could not be updated.")
    print("Manual verification may be needed for:")
    for extracted_name, excel_name, _ in matched_files:
        if excel_name is None:
            print(f"  - {extracted_name}")

print("\n" + "=" * 80)
print("EXCEL UPDATE COMPLETE")
print("=" * 80)
print(f"\nFile: {excel_file}")
print("Please review the updated spreadsheet for accuracy.")
print("=" * 80)
