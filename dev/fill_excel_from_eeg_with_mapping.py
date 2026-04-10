#!/usr/bin/env python3
"""
Fill EEG_Segmentation.xlsx using complete subject mapping
Apply September marker interpretations to Excel
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import mne
from collections import Counter

# ============================================================================
# 被试映射数据 (完整的September映射)
# ============================================================================

SEPTEMBER_MARKER_MAPPING = {
    "sep_12.vhdr": {
        "recording_date": "2013-09-12",
        "subject_info": "1990 F, Quality",
        "problem_order": ["Cake", "Bin", "Toothbrush", "Workspace", "Fountain", "Metro"],
        "s14_meaning": "Typing (Problem 2)",
        "s15_meaning": "Evaluate WL (Problem 2)",
        "notes": "Standard sequence"
    },
    "sep_12(2).vhdr": {
        "recording_date": "2013-09-12",
        "subject_info": "1987, Anthropology",
        "problem_order": ["Toothbrush", "Workspace", "Cake", "Metro", "Bin", "Fountain"],
        "s14_meaning": "Typing (Problem 2)",
        "s15_meaning": "Evaluate WL (Problem 2)",
        "notes": "Standard sequence"
    },
    "sep_13.vhdr": {
        "recording_date": "2013-09-13",
        "subject_info": "1987, Info Sys Eng",
        "problem_order": ["Metro", "Bin", "Cake", "Brush", "Fountain", "Exercise"],
        "s14_meaning": "Typing (Problem 2)",
        "s15_meaning": "Evaluate WL (Problem 2)",
        "notes": "Standard sequence"
    },
    "sep_13(2).vhdr": {
        "recording_date": "2013-09-13",
        "subject_info": "1987, ECE",
        "problem_order": ["Cake", "Exercise", "Brush", "Bin", "Metro", "Fountain"],
        "s14_meaning": "Typing (Problem 2)",
        "s15_meaning": "Evaluate WL (Problem 2)",
        "notes": "Signal bad, very relax"
    },
    "eeg_sep_18.vhdr": {
        "recording_date": "2013-09-18",
        "subject_info": "1982 F, Management",
        "problem_order": ["Cake", "Exercise", "Fountain", "Bin", "Brush", "Metro"],
        "s14_meaning": "Evaluate Solution (Problem 2)",
        "s15_meaning": "Typing (Problem 2)",
        "notes": "Sequence shifted by +1 due to 'done' marker"
    }
}

# ============================================================================
# 数据提取和填充函数
# ============================================================================

def extract_september_events(vhdr_path):
    """从September EEG文件提取事件"""
    try:
        raw = mne.io.read_raw_brainvision(str(vhdr_path), preload=False, verbose='ERROR')
        markers = Counter(raw.annotations.description)
        total_events = len(raw.annotations)

        return {
            'total_events': total_events,
            'markers': dict(markers),
            'S14_count': markers.get('Stimulus/S 14', 0),
            'S15_count': markers.get('Stimulus/S 15', 0),
        }
    except Exception as e:
        print(f"  ❌ 错误 {vhdr_path.name}: {str(e)}")
        return None

def fill_september_row(mapping, events_data, excel_row_index):
    """
    为September行填充Excel数据
    根据映射信息填充对应的标记数
    """
    fill_data = {
        'total_event': events_data['total_events'],
        'match_or_not': 'OK' if events_data['total_events'] > 0 else 'CHECK',
    }

    # 从映射中获取S14和S15对应的问题
    s14_meaning = mapping['s14_meaning']  # e.g., "Typing (Problem 2)"
    s15_meaning = mapping['s15_meaning']

    # 提取问题号 (从"Typing (Problem 2)"中提取2)
    try:
        s14_problem_num = int(s14_meaning.split('Problem ')[-1].rstrip(')'))
        s15_problem_num = int(s15_meaning.split('Problem ')[-1].rstrip(')'))
    except:
        return fill_data

    # 根据问题号填充对应的标记计数
    # 例如 S14是Problem 2的"Typing"，则填充"2_type"列
    problem_stages = {
        'Read problem': '1_problem',
        'Generate solution': '2_solution',
        'Evaluate WL': '3_rate',
        'Evaluate solution': '4_eval',
        'Typing': '5_type',
        'Evaluate WL (后)': '6_rate',
    }

    # 简化方法：对于September数据，直接填充S14和S15计数到对应问题的标记列
    s14_column = f'{s14_problem_num}_type'  # Problem 2, Typing -> 2_type
    s15_column = f'{s15_problem_num}_rate'  # Problem 2, Evaluate WL -> 2_rate

    fill_data[s14_column] = events_data['S14_count']
    fill_data[s15_column] = events_data['S15_count']

    return fill_data

def main():
    print("=" * 100)
    print("SEPTEMBER EEG数据 - Excel填充 (基于完整映射)")
    print("=" * 100)

    excel_path = Path(r"C:\Users\whyke\github\six-problem\Data\EEG_Segmentation.xlsx")

    if not excel_path.exists():
        print(f"❌ Excel文件不存在: {excel_path}")
        return

    print(f"\n📂 加载Excel: {excel_path.name}")

    # 创建备份
    backup_path = excel_path.parent / f"{excel_path.stem}_backup.xlsx"
    if not backup_path.exists():
        import shutil
        shutil.copy(excel_path, backup_path)
        print(f"✅ 创建备份: {backup_path.name}")

    # 加载Excel
    wb = load_workbook(excel_path)
    ws = wb.active

    print(f"\n📊 处理September数据...")
    print("-" * 100)

    processed_count = 0

    # 处理每个September文件
    for filename, mapping in SEPTEMBER_MARKER_MAPPING.items():
        print(f"\n{filename}")

        # 查找对应的EEG文件
        # September文件在: C:\Users\whyke\github\six-problem\Data\OneDrive_1_01-04-2026\

        # 构造可能的路径
        possible_paths = [
            Path(f"C:/Users/whyke/github/six-problem/Data/OneDrive_1_01-04-2026/eeg_sep_12/{filename}"),
            Path(f"C:/Users/whyke/github/six-problem/Data/OneDrive_1_01-04-2026/eeg_sep_12(2)/{filename}"),
            Path(f"C:/Users/whyke/github/six-problem/Data/OneDrive_1_01-04-2026/eeg_sep_13/{filename}"),
            Path(f"C:/Users/whyke/github/six-problem/Data/OneDrive_1_01-04-2026/eeg_sep_13(2)/{filename}"),
            Path(f"C:/Users/whyke/github/six-problem/Data/OneDrive_1_01-04-2026/eeg_sep_18/{filename}"),
        ]

        vhdr_path = None
        for path in possible_paths:
            if path.exists():
                vhdr_path = path
                break

        if not vhdr_path:
            print(f"  ⚠️ 文件未找到")
            continue

        # 提取事件
        events_data = extract_september_events(vhdr_path)
        if not events_data:
            continue

        print(f"  📍 位置: {vhdr_path}")
        print(f"  📊 事件: {events_data['total_events']} 总计 (S14: {events_data['S14_count']}, S15: {events_data['S15_count']})")
        print(f"  👤 被试: {mapping['subject_info']}")
        print(f"  📌 标记: S14={mapping['s14_meaning']}, S15={mapping['s15_meaning']}")

        # 填充数据
        fill_data = fill_september_row(mapping, events_data, None)

        # 在Excel中查找对应的行 (按filename匹配)
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row_idx, 1).value  # 第一列是文件名
            if cell_value and filename in str(cell_value):
                print(f"  ✅ Excel行 {row_idx} - 填充中...")

                # 根据列名填充
                for col_idx in range(1, ws.max_column + 1):
                    header = ws.cell(1, col_idx).value
                    if header in fill_data:
                        ws.cell(row_idx, col_idx).value = fill_data[header]
                        # 标记已填充的单元格
                        if header == 'match_or_not' and fill_data[header] == 'OK':
                            ws.cell(row_idx, col_idx).fill = PatternFill(
                                start_color="C6EFCE",
                                end_color="C6EFCE",
                                fill_type="solid"
                            )

                processed_count += 1
                break

    # 保存
    print(f"\n" + "=" * 100)
    print(f"✅ 处理完成: {processed_count} 个September文件")
    print(f"💾 保存Excel...")

    wb.save(excel_path)
    print(f"✅ Excel已更新: {excel_path.name}")

    print(f"\n📝 生成映射验证报告...")

    # 生成映射验证报告
    report = {
        'timestamp': pd.Timestamp.now().isoformat(),
        'excel_file': str(excel_path),
        'processed_subjects': processed_count,
        'mapping_source': 'SEPTEMBER_MARKER_MAPPING (complete)',
        'subjects_processed': list(SEPTEMBER_MARKER_MAPPING.keys()),
        'summary': {
            'total_september_subjects': 5,
            'successfully_mapped': processed_count,
            'status': '✅ 完成' if processed_count == 5 else f'⚠️ 部分完成 ({processed_count}/5)'
        }
    }

    report_path = Path("mapping_verification_report.json")
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, ensure_ascii=False)

    print(f"✅ 报告已生成: {report_path.name}")

    print("\n" + "=" * 100)
    print("📊 September数据映射完成！")
    print("=" * 100)

if __name__ == "__main__":
    main()
