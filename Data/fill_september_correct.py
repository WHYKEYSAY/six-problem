# -*- coding: utf-8 -*-
"""
Fill September rows in EEG_Segmentation.xlsx based on NOTEBOOK images.

Method:
- Notebook gives sequential marker numbers (2, 3, 4, ...) for each event type
- EEG confirms total event count (all markers present, plus 1 extra pre-experiment)
- Fill each cell with the notebook marker number
- Leave blank (None) where notebook says event is missing

Structure per row:
  Col D = eye_closed
  Col E-J = P1: prob, sol, rate(WL pre), eval, type, rate(WL post)
  Col K-P = P2, Col Q-V = P3, Col W-AB = P4, Col AC-AH = P5, Col AI-AN = P6
  Col AO = eye_closed_2
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path

EXCEL_PATH = Path(r"C:\Users\whyke\github\six-problem\Data\EEG_Segmentation.xlsx")

# ============================================================
# NOTEBOOK DATA (from handwritten notebook images)
# Each row: [eye_closed, p1×6, p2×6, p3×6, p4×6, p5×6, p6×6, eye_closed_2]
# None = marker missing (leave blank)
# ============================================================

SEPTEMBER_ROWS = {

    # Row 21: Sep_12(2).xlsx → Sep 12, 2013 (2) Anthropology background
    # Notebook: 2(eye close), 3(stop), 4-9(P1), 10-15(P2), 16-21(P3),
    #           22-26(P4, note: "miss typing or WL?" → typing blank),
    #           27-32(P5), 33-38(P6), 39(end), 40-41(eye closed2)
    # EEG total: 41 events ✓ (marker 1 extra + markers 2-41 = 41)
    21: {
        "name": "Sep_12(2).xlsx",
        "total_event": 41,
        "markers": [
            2,           # eye_closed
            4, 5, 6, 7, 8, 9,          # P1
            10, 11, 12, 13, 14, 15,     # P2
            16, 17, 18, 19, 20, 21,     # P3
            22, 23, 24, 25, None, 26,   # P4 (typing missing per notebook note)
            27, 28, 29, 30, 31, 32,     # P5
            33, 34, 35, 36, 37, 38,     # P6 (39=end extra, not in table)
            40,          # eye_closed_2
        ]
    },

    # Row 22: Sep_12.xlsx → Sep 12, 2013 (1)
    # Notebook: 2(eye close), 3(stop), 4-9(P1), 10-15(P2),
    #           16-22(P3 + 18=draw extra → skip 18 in table),
    #           23-28(P4 "change mind/again"), 29-34(P5),
    #           35-41(P6 + 41=stop extra → skip 41 in table), 42-43(eye closed2)
    # EEG total: 43 events ✓
    22: {
        "name": "Sep_12.xlsx",
        "total_event": 43,
        "markers": [
            2,           # eye_closed
            4, 5, 6, 7, 8, 9,          # P1
            10, 11, 12, 13, 14, 15,    # P2
            16, 17, 19, 20, 21, 22,    # P3 (18=draw extra, skip → rate starts at 19)
            23, 24, 25, 26, 27, 28,    # P4
            29, 30, 31, 32, 33, 34,    # P5
            35, 36, 37, 38, 39, 40,    # P6 (41=stop extra, skip)
            42,          # eye_closed_2
        ]
    },

    # Row 23: Sep_13(2).xlsx → Sep 13, 2013 (2) ECE
    # Notebook: 2(eye close), 3, 4-9(P1), 10-15(P2),
    #           16-22(P3 + 22=done extra), 23-28(P4),
    #           29-34(P5), 35-41(P6 + 41=done extra), 42-43(eye closed2)
    # EEG total: 43 events ✓
    23: {
        "name": "Sep_13(2).xlsx",
        "total_event": 43,
        "markers": [
            2,           # eye_closed
            4, 5, 6, 7, 8, 9,          # P1
            10, 11, 12, 13, 14, 15,    # P2
            16, 17, 18, 19, 20, 21,    # P3 (22=done extra, not in table)
            23, 24, 25, 26, 27, 28,    # P4 (starts at 23 after done=22)
            29, 30, 31, 32, 33, 34,    # P5
            35, 36, 37, 38, 39, 40,    # P6 (41=done extra, not in table)
            42,          # eye_closed_2
        ]
    },

    # Row 24: Sep_13.xlsx → Sep 13, 2013 (1) Info & Sys Eng
    # Notebook: 2(eye close), 3, 4-9(P1), 10-15(P2), 16-21(P3),
    #           22-27(P4), 28-33(P5), 34-40(P6 + 40=done extra), 41(eye closed2)
    # EEG total: 42 events ✓
    24: {
        "name": "Sep_13.xlsx",
        "total_event": 42,
        "markers": [
            2,           # eye_closed
            4, 5, 6, 7, 8, 9,          # P1
            10, 11, 12, 13, 14, 15,    # P2
            16, 17, 18, 19, 20, 21,    # P3
            22, 23, 24, 25, 26, 27,    # P4
            28, 29, 30, 31, 32, 33,    # P5
            34, 35, 36, 37, 38, 39,    # P6 (40=done extra, not in table)
            41,          # eye_closed_2
        ]
    },

    # Row 25: Sep_18.xlsx → Sep 18, 2013 (Management)
    # Notebook: 2(eye close), 3, 4-10(P1 + 10=done extra), 11-16(P2),
    #           17-22(P3), 23-28(P4), 29-34(P5),
    #           35-41(P6 + 41=done extra), 42-43(eye closed2)
    # EEG total: 44 events ✓
    25: {
        "name": "Sep_18.xlsx",
        "total_event": 44,
        "markers": [
            2,           # eye_closed
            4, 5, 6, 7, 8, 9,          # P1 (10=done extra, not in table)
            11, 12, 13, 14, 15, 16,    # P2
            17, 18, 19, 20, 21, 22,    # P3
            23, 24, 25, 26, 27, 28,    # P4
            29, 30, 31, 32, 33, 34,    # P5
            35, 36, 37, 38, 39, 40,    # P6 (41=done extra, not in table)
            42,          # eye_closed_2
        ]
    },
}

COLUMN_LABELS = [
    "eye_closed",
    "1_prob", "1_sol", "1_rate", "1_eval", "1_type", "1_rate2",
    "2_prob", "2_sol", "2_rate", "2_eval", "2_type", "2_rate2",
    "3_prob", "3_sol", "3_rate", "3_eval", "3_type", "3_rate2",
    "4_prob", "4_sol", "4_rate", "4_eval", "4_type", "4_rate2",
    "5_prob", "5_sol", "5_rate", "5_eval", "5_type", "5_rate2",
    "6_prob", "6_sol", "6_rate", "6_eval", "6_type", "6_rate2",
    "eye_closed_2",
]  # 38 values total (col D to col AO)


def get_col_letter(col_num):
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(ord('A') + col_num % 26) + result
        col_num //= 26
    return result


def fill_september():
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    print("=" * 90)
    print("FILLING SEPTEMBER ROWS (21-25) BASED ON NOTEBOOK DATA")
    print("=" * 90)

    for row_num, data in SEPTEMBER_ROWS.items():
        name = data["name"]
        total = data["total_event"]
        markers = data["markers"]

        # Verify length
        assert len(markers) == 38, f"{name}: expected 38 marker values, got {len(markers)}"

        # Write Name, total_event, match_or_not
        ws[f'A{row_num}'] = name
        ws[f'B{row_num}'] = total
        ws[f'C{row_num}'] = "Match"

        # Clear old consecutive data first (cols D to AO = cols 4 to 41)
        for col in range(4, 42):
            ws.cell(row_num, col).value = None

        # Write new marker data starting at col D (col 4)
        missing = []
        for i, val in enumerate(markers):
            col_num = 4 + i   # D=4, E=5, ...
            ws.cell(row_num, col_num).value = val
            if val is None:
                missing.append(COLUMN_LABELS[i])

        # Summary
        filled = [v for v in markers if v is not None]
        print(f"\nRow {row_num}: {name}")
        print(f"  total_event: {total} | Markers filled: {len(filled)}/38")
        print(f"  Range: {filled[0]} to {filled[-1]}")
        if missing:
            print(f"  ⚠ Missing (blank): {missing}")
        else:
            print(f"  ✓ All 38 cells filled")
        print(f"  Sequence: {markers}")

    wb.save(EXCEL_PATH)
    print(f"\n✓ Saved: {EXCEL_PATH}")


if __name__ == "__main__":
    fill_september()

    # Print verification table
    print("\n" + "=" * 90)
    print("VERIFICATION SUMMARY")
    print("=" * 90)
    print(f"{'Row':<5} {'File':<22} {'EEG events':>10} {'In table':>9} {'Eye start':>10} {'Eye end':>8} {'P1':>6} {'P2':>6} {'P3 note'}")
    print("-" * 90)
    for row, d in SEPTEMBER_ROWS.items():
        m = d["markers"]
        filled = [v for v in m if v is not None]
        p1 = f"{m[1]}-{m[6]}"
        p2 = f"{m[7]}-{m[12]}"
        note = ""
        if None in m:
            idx = m.index(None)
            note = f"blank at col {COLUMN_LABELS[idx]}"
        print(f"{row:<5} {d['name']:<22} {d['total_event']:>10} {len(filled):>9} {str(m[0]):>10} {str(m[-1]):>8} {p1:>6} {p2:>6}  {note}")
