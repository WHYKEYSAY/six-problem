# -*- coding: utf-8 -*-
"""
Fill EEG_Segmentation.xlsx with September data based on actual EEG extraction results.
Marker pattern: consecutive numbers starting from 4
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook
from pathlib import Path

# September extraction results (from extract_september_simple.py)
SEPTEMBER_DATA = {
    "Sep_12(2).xlsx": {
        "total_event": 41,
        "match_or_not": "Match"
    },
    "Sep_12.xlsx": {
        "total_event": 43,
        "match_or_not": "Match"
    },
    "Sep_13(2).xlsx": {
        "total_event": 43,
        "match_or_not": "Match"
    },
    "Sep_13.xlsx": {
        "total_event": 42,
        "match_or_not": "Match"
    },
    "Sep_18.xlsx": {
        "total_event": 44,
        "match_or_not": "Match"
    }
}

def generate_marker_sequence(start=4, count=None):
    """Generate consecutive marker numbers from start"""
    if count is None:
        return []
    return list(range(start, start + count))

def get_column_letter(col_num):
    """Convert column number to letter(s): 1->A, 2->B, 27->AA, etc."""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(ord('A') + col_num % 26) + result
        col_num //= 26
    return result

def fill_excel():
    excel_path = Path(r"C:\Users\whyke\github\six-problem\Data\EEG_Segmentation.xlsx")

    print(f"Opening {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb.active

    # Excel rows 21-25 for the 5 September files
    start_row = 21

    for idx, (filename, data) in enumerate(SEPTEMBER_DATA.items()):
        row = start_row + idx

        # Column A: Name
        ws[f'A{row}'] = filename

        # Column B: total_event
        ws[f'B{row}'] = data['total_event']

        # Column C: match_or_not
        ws[f'C{row}'] = data['match_or_not']

        # Columns D onwards: marker numbers (starting from column D, which is column 4)
        markers = generate_marker_sequence(start=4, count=data['total_event'])

        for col_idx, marker in enumerate(markers):
            # Column D = 4, E = 5, etc.
            col_num = 4 + col_idx
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}{row}'] = marker

        print(f"\n✓ Row {row}: {filename}")
        print(f"  total_event: {data['total_event']}")
        print(f"  markers: 4-{4 + data['total_event'] - 1} ({data['total_event']} markers)")

    # Save
    wb.save(excel_path)
    print(f"\n✓ File saved: {excel_path}")

if __name__ == "__main__":
    print("="*80)
    print("FILLING SEPTEMBER DATA INTO EEG_Segmentation.xlsx")
    print("="*80)

    fill_excel()

    print("\n" + "="*80)
    print("SUMMARY OF SEPTEMBER MARKER SEQUENCES")
    print("="*80)

    for filename, data in SEPTEMBER_DATA.items():
        total = data['total_event']
        markers_start = 4
        markers_end = 4 + total - 1
        print(f"\n{filename}:")
        print(f"  Total events: {total}")
        print(f"  Marker range: {markers_start}-{markers_end}")
        print(f"  Formula: 4, 5, 6, ..., {markers_end}")
