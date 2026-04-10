# -*- coding: utf-8 -*-
"""
Verify that EEG_Segmentation.xlsx was successfully filled with September data
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from openpyxl import load_workbook
from pathlib import Path

def verify_excel():
    excel_path = Path(r"C:\Users\whyke\github\six-problem\Data\EEG_Segmentation.xlsx")

    print("="*100)
    print("VERIFYING EEG_Segmentation.xlsx - SEPTEMBER ROWS (21-25)")
    print("="*100)

    wb = load_workbook(excel_path)
    ws = wb.active

    for row in range(21, 26):
        name = ws[f'A{row}'].value
        total_event = ws[f'B{row}'].value
        match_status = ws[f'C{row}'].value

        # Get marker columns (D onwards)
        markers = []
        col_idx = 4  # Column D
        while True:
            cell_value = ws.cell(row, col_idx).value
            if cell_value is None:
                break
            markers.append(cell_value)
            col_idx += 1

        print(f"\n【Row {row}】{name}")
        print(f"  total_event: {total_event}")
        print(f"  match_or_not: {match_status}")

        if markers:
            print(f"  Markers: {markers[0]}, {markers[1]}, {markers[2]}, ..., {markers[-1]}")
            print(f"  Count: {len(markers)} markers")

            # Verify they are consecutive
            expected_range = range(4, 4 + total_event)
            if list(markers) == list(expected_range):
                print(f"  ✓ VERIFIED: Consecutive from 4 to {4 + total_event - 1}")
            else:
                print(f"  ✗ ERROR: Markers don't match expected range")
        else:
            print(f"  ⚠ WARNING: No markers found in row {row}")

    print("\n" + "="*100)
    print("VERIFICATION COMPLETE")
    print("="*100)

if __name__ == "__main__":
    verify_excel()
